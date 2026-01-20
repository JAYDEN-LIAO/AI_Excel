# backend/formula_service.py
import openpyxl
from openpyxl.utils import column_index_from_string
import os
import uuid
import pandas as pd
import excel_ops

def apply_formula_to_file(file_path: str, ai_result: dict):
    # 1. 基础文件准备
    dir_name = os.path.dirname(file_path)
    _, ext = os.path.splitext(os.path.basename(file_path))
    safe_name = f"processed_{uuid.uuid4().hex[:8]}{ext}"
    new_file_path = os.path.join(dir_name, safe_name)

    # 2. 读取数据
    try:
        df = pd.read_excel(file_path)
    except Exception as e:
        raise Exception(f"无法读取 Excel 文件: {e}")

    # 3. 准备 Python 执行沙箱
    safe_env = {
        "excel_ops": excel_ops,
        "df": df,
        "pd": pd,
        "rows": df.to_dict('records'),
        "row": None
    }

    action_type = ai_result.get('action_type', 'formula')
    py_expr = ai_result.get('python_expression', '')
    mode = ai_result.get('mode', 'column')
    # AI 可能返回 "性别" (列名) 也可能返回 "B" (列字母)
    target_pos = str(ai_result.get('target_position', 'AI计算结果')).strip()

    print(f"--- 🚀 开始执行: {action_type} (模式: {mode}) ---")
    print(f"--- 🐍 Python 执行代码: {py_expr} ---")
    print(f"--- 🎯 目标位置: {target_pos} ---")

    try:
        # ==========================================
        # 🟢 分支 A: 结构修改 (Structure Mode)
        # ==========================================
        if action_type == 'structure':
            try:
                exec(py_expr, globals(), safe_env)
            except Exception as e:
                raise Exception(f"结构修改代码执行失败: {e}")
            new_df = safe_env.get('df')
            if new_df is None: raise ValueError("DataFrame 丢失")
            new_df.to_excel(new_file_path, index=False)
            return new_file_path, safe_name

        # ==========================================
        # 🔵 分支 B: 值计算模式 (Value Calculation Mode)
        # ==========================================
        else:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active

            if mode == 'column':
                # --- 1. Python 计算逻辑 (保持不变) ---
                calculated_values = []
                def calc_single_row(current_row):
                    safe_env['row'] = current_row
                    try:
                        return eval(py_expr, globals(), safe_env)
                    except Exception as e:
                        return f"Error: {str(e)}"

                print("⏳ 正在进行 Python 内存计算...")
                series_result = df.apply(calc_single_row, axis=1)
                calculated_values = series_result.tolist()

                # --- 2. 智能定位逻辑 (🔥 核心逻辑升级 🔥) ---
                print(f"💾 正在定位目标列 '{target_pos}'...")
                target_col_idx = None

                # 📌 策略 A: 优先匹配现有的【表头名称】
                # (解决：用户说“写入年龄列”，直接覆盖原“年龄”列)
                for cell in ws[1]:
                    # 强转 string 比较，忽略空格
                    if str(cell.value).strip() == target_pos:
                        target_col_idx = cell.column
                        print(f"✅ 按表头名匹配成功: '{target_pos}' -> 第 {target_col_idx} 列")
                        break

                # 📌 策略 B: 如果没找到表头，尝试解析为【Excel 列字母】(如 "G", "AA")
                # (解决：用户说“写入 G 列”，即使 G 列目前是空的，也要定位到第 7 列)
                if target_col_idx is None:
                    # 只有当它是纯字母，且长度合理(<=3)时才认为是列标 (避免把 "Total" 误判为 T列)
                    if target_pos.isalpha() and len(target_pos) <= 3:
                        try:
                            # 强制转换为大写并获取索引 (例如 "G" -> 7)
                            potential_idx = column_index_from_string(target_pos.upper())

                            # 🔥 核心修改：只要是合法的正整数列号，就直接采纳！
                            # 不再检查 <= ws.max_column，允许跳跃写入。
                            if potential_idx > 0:
                                target_col_idx = potential_idx
                                print(f"📍 按列坐标定位: '{target_pos}' -> 第 {target_col_idx} 列")

                                # 💡 细节优化：如果这一列还没有表头，把 target_pos 填进去
                                # 比如跳到 G 列，G1 是空的，就填入 "G"
                                header_cell = ws.cell(row=1, column=target_col_idx)
                                if not header_cell.value:
                                    header_cell.value = target_pos
                                    print(f"📝 自动补充表头: {target_pos}")
                        except:
                            pass # 转换失败（说明不是列字母），继续往下走

                # 📌 策略 C: 既不是现有表头，也不是列字母，说明是【完全的新列名】
                # (解决：用户说“写入新列[预测值]”，则追加到最后)
                if target_col_idx is None:
                    target_col_idx = ws.max_column + 1
                    ws.cell(row=1, column=target_col_idx, value=target_pos) # 写表头
                    print(f"🆕 目标是新字段，追加到末尾: '{target_pos}' -> 第 {target_col_idx} 列")

                # --- 3. 写入数据 ---
                for i, val in enumerate(calculated_values):
                    # i 从 0 开始，Excel 数据从第 2 行开始
                    excel_row = i + 2

                    # 即使跳到了第 7 列，行号依然受 max_row 限制 (不要写到无限行)
                    if excel_row <= ws.max_row:
                        cell = ws.cell(row=excel_row, column=target_col_idx)
                        try:
                            if hasattr(val, 'item'): val = val.item()
                            cell.value = val
                        except:
                            cell.value = str(val)

            # 单元格模式 (保持不变)
            elif mode == 'cell':
                try:
                    final_value = eval(py_expr, globals(), safe_env)
                except Exception as e:
                    raise Exception(f"聚合计算失败: {e}")

                if target_pos:
                    ws[target_pos] = final_value
                else:
                    print("⚠️ 未指定 target_position")

            wb.save(new_file_path)
            print(f"✅ 处理完成: {safe_name}")
            return new_file_path, safe_name

    except Exception as e:
        print(f"❌ 严重错误: {e}")
        raise e

# formula_service.py (追加)

def apply_multi_file_operation(file_map: dict, py_code: str):
    """
    执行多表关联操作
    :param file_map: { "文件名": "物理路径" }
    :param py_code: AI 生成的 Python 代码
    """
    import pandas as pd
    import uuid
    import os

    # 1. 准备环境：加载所有 DataFrame
    dfs = {}
    print("--- 🔄 正在加载多表上下文 ---")
    try:
        for fname, fpath in file_map.items():
            # 简单起见，默认读第一个 Sheet
            dfs[fname] = pd.read_excel(fpath)
            print(f"✅ 已加载: {fname} ({len(dfs[fname])} 行)")
    except Exception as e:
        raise Exception(f"加载文件失败: {fname} -> {e}")

    # 2. 准备沙箱
    safe_env = {
        "pd": pd,
        "dfs": dfs,
        "result_df": None # 占位符
    }

    # 3. 执行 AI 代码
    print(f"--- 🐍 执行多表代码 ---\n{py_code}")
    try:
        exec(py_code, globals(), safe_env)
    except Exception as e:
        raise Exception(f"代码执行错误: {e}")

    # 4. 获取结果
    final_df = safe_env.get('result_df')
    if final_df is None or not isinstance(final_df, pd.DataFrame):
        raise Exception("代码执行完毕，但 `result_df` 变量为空或不是 DataFrame")

    # 5. 保存结果到新文件
    # 取第一个文件的目录作为输出目录
    first_path = list(file_map.values())[0]
    dir_name = os.path.dirname(first_path)
    safe_name = f"multi_result_{uuid.uuid4().hex[:8]}.xlsx"
    output_path = os.path.join(dir_name, safe_name)

    final_df.to_excel(output_path, index=False)
    print(f"✅ 多表处理完成，已保存至: {output_path}")

    return output_path, safe_name